from openpyxl import load_workbook, Workbook
import time
from datetime import datetime
import psutil
import os

def get_memory_usage():
    process = psutil.Process(os.getpid())
    return process.memory_info().rss / 1024 / 1024  # Converte para MB

def log_memory_and_time(message, start_time, start_memory):
    current_time = time.time()
    current_memory = get_memory_usage()
    elapsed_time = current_time - start_time
    memory_used = current_memory - start_memory
    log_message = f"{message} - Tempo: {elapsed_time:.2f} segundos, Memória usada: {memory_used:.2f} MB"
    print(log_message)
    return log_message

# Configurações considerando um arquivo excel com 800 mil linhas
input_file = '800k.xlsx'
output_file = f'output_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
log_file = f'log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'

# Inicialização
start_time = time.time()
start_memory = get_memory_usage()

# Fase 1: Leitura do Excel
print("Iniciando leitura do arquivo Excel...")
wb = load_workbook(filename=input_file, read_only=True)
ws = wb.active

data = []
for row in ws.iter_rows(values_only=True):
    # Assumindo que queremos as colunas 5 (NOME) e 11 (Valor)
    # Ajuste os índices se necessário
    data.append(([row[2], row[3], row[10]]))

wb.close()

# Log após a leitura
log_after_read = log_memory_and_time("Leitura concluída", start_time, start_memory)

# Fase 2: Escrita para o novo Excel
print("Iniciando escrita para o novo arquivo Excel...")
new_wb = Workbook(write_only=True)
new_ws = new_wb.create_sheet()

# Escreve o cabeçalho
new_ws.append(["NOME", "VALOR"])

# Escreve os dados
for row in data:
    new_ws.append(row)

new_wb.save(output_file)
new_wb.close()

# Log após a escrita
log_after_write = log_memory_and_time("Escrita concluída", start_time, start_memory)

# Limpeza de memória
del data
del wb
del new_wb

# Log final
final_log = log_memory_and_time("Processo finalizado", start_time, start_memory)

# Salva os logs em um arquivo
with open(log_file, 'w') as f:
    f.write(log_after_read + '\n')
    f.write(log_after_write + '\n')
    f.write(final_log + '\n')

print(f"Processo concluído. Novo arquivo Excel salvo como {output_file}")
print(f"Log salvo em {log_file}")
