from openpyxl import load_workbook, Workbook
import time
from datetime import datetime
import psutil
import os

def get_memory_usage():
    process = psutil.Process(os.getpid())
    return process.memory_info().rss / 1024 / 1024  # Converte para MB

def process_row(row):
    # Assumindo que 'NOME' é a 5ª coluna (índice 4) e 'Valor da M/E' é a 11ª coluna (índice 10)
    return [row[2].value, row[3].value, row[10].value]

def is_row_empty(row):
    return all(cell.value is None for cell in row)

# Cria nomes de arquivo únicos baseados no timestamp
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_filename = f"processamento_{timestamp}.txt"
new_excel_filename = f"novo_excel_{timestamp}.xlsx"

# Registra o tempo de início e uso de memória inicial
start_time = time.time()
initial_memory = get_memory_usage()

print(f"Abrindo o arquivo Excel... Dados de processamento serão salvos em {output_filename}")
print(f"Novo arquivo Excel será salvo como {new_excel_filename}")

wb = load_workbook(filename='800k.xlsx', read_only=True)
ws = wb.active

# Cria um novo workbook para o output
new_wb = Workbook(write_only=True)
new_ws = new_wb.create_sheet()

print("Processando linhas...")
row_count = 0
empty_row_count = 0
max_empty_rows = 10  # Número máximo de linhas vazias consecutivas antes de parar

with open(output_filename, 'w', encoding='utf-8') as output_file:
    output_file.write("Início do processamento\n")

    # Escreve o cabeçalho no novo arquivo Excel
    new_ws.append(["CPF", "Nome" , "Valor da M/E"])

    for row in ws.iter_rows():
        if is_row_empty(row):
            empty_row_count += 1
            if empty_row_count >= max_empty_rows:
                msg = f"Encontradas {max_empty_rows} linhas vazias consecutivas. Parando o processamento."
                print(msg)
                output_file.write(msg + "\n")
                break
        else:
            empty_row_count = 0  # Reseta o contador se encontrar uma linha não vazia
            processed_row = process_row(row)
            
            # Adiciona a linha processada ao novo arquivo Excel
            new_ws.append(processed_row)
            
            row_count += 1
            
            # Registra o progresso a cada 10000 linhas
            if row_count % 10000 == 0:
                current_time = time.time()
                elapsed_time = current_time - start_time
                current_memory = get_memory_usage()
                memory_used = current_memory - initial_memory
                progress_msg = f"Processadas {row_count} linhas... Tempo: {elapsed_time:.2f} segundos, Memória usada: {memory_used:.2f} MB"
                output_file.write(progress_msg + "\n")
                output_file.flush()  # Força a escrita no arquivo
                print(progress_msg)

    # Salva o novo arquivo Excel
    new_wb.save(new_excel_filename)

    # Calcula o tempo total e uso de memória final
    end_time = time.time()
    total_time = end_time - start_time
    final_memory = get_memory_usage()
    total_memory_used = final_memory - initial_memory

    summary = f"\nProcessamento concluído.\n"
    summary += f"Total de linhas processadas: {row_count}\n"
    summary += f"Tempo total de processamento: {total_time:.2f} segundos\n"
    summary += f"Memória total utilizada: {total_memory_used:.2f} MB\n"

    print(summary)
    output_file.write(summary)

print(f"Processamento concluído. Dados de processamento salvos em {output_filename}")
print(f"Novo arquivo Excel salvo como {new_excel_filename}")

# Mede a memória antes da desalocação
memory_before_deallocation = get_memory_usage()

# Libera as referências aos objetos grandes
wb.close()
new_wb.close()
del wb
del new_wb
del ws
del new_ws

# Mede a memória após a desalocação
memory_after_deallocation = get_memory_usage()
memory_deallocated = memory_before_deallocation - memory_after_deallocation

print("\nInformações sobre desalocação de memória:")
print(f"Memória antes da desalocação: {memory_before_deallocation:.2f} MB")
print(f"Memória após a desalocação: {memory_after_deallocation:.2f} MB")
print(f"Memória desalocada: {memory_deallocated:.2f} MB")

# Adiciona estas informações ao arquivo de saída
with open(output_filename, 'a', encoding='utf-8') as output_file:
    output_file.write("\nInformações sobre desalocação de memória:\n")
    output_file.write(f"Memória antes da desalocação: {memory_before_deallocation:.2f} MB\n")
    output_file.write(f"Memória após a desalocação: {memory_after_deallocation:.2f} MB\n")
    output_file.write(f"Memória desalocada: {memory_deallocated:.2f} MB\n")
